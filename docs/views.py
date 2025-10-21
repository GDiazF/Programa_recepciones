from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
import openpyxl
import pandas as pd
from io import BytesIO
from django.contrib import messages
from django.shortcuts import redirect
from django.db import models
from .models import Directores, Establecimientos, Proveedor, TipoProveedor, TipoRecibo, Servicios, RegistroServicio
import time
from django.db import transaction
from datetime import datetime
from django.core.paginator import Paginator
from .forms import ServicioForm, EstablecimientoForm, ProveedorForm, TipoReciboForm, RegistroServicioForm, PerfilUsuarioForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import PasswordChangeView, LoginView
from django.contrib.auth.forms import PasswordChangeForm
from django.urls import reverse_lazy

#reportlab import
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.colors import HexColor
import io
from django.conf import settings
import os
from calendar import month_name
import locale
from pathlib import Path
from django.templatetags.static import static
from django.urls import reverse
from urllib.parse import quote
import base64
from django.contrib.auth.mixins import LoginRequiredMixin

# Create your views here.

class CustomLoginView(LoginView):
    template_name = 'docs/login.html'
    redirect_authenticated_user = True
    
    def get_success_url(self):
        return reverse_lazy('base')

#vista para la pagina principal - redirige al dashboard
@login_required
def base(request):
    return redirect('docs:dashboard')

#vista para el dashboard con estadísticas
@login_required
def dashboard(request):
    from django.db.models import Count
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    # Fecha actual
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    
    # Estadísticas generales
    total_registros = RegistroServicio.objects.count()
    total_servicios = Servicios.objects.count()
    total_establecimientos = Establecimientos.objects.count()
    total_proveedores = Proveedor.objects.count()
    
    # Registros del mes actual
    cantidad_mes = RegistroServicio.objects.filter(
        fecha_emision__gte=inicio_mes,
        fecha_emision__lte=hoy
    ).count()
    
    # Obtener filtros
    proveedor_id = request.GET.get('proveedor')
    establecimiento_id = request.GET.get('establecimiento')
    servicio_id = request.GET.get('servicio')
    
    # Query base para últimos registros
    ultimos_registros = RegistroServicio.objects.select_related(
        'servicio',
        'servicio__establecimiento',
        'servicio__proveedor',
        'servicio__tipo_recibo'
    )
    
    # Aplicar filtros
    if proveedor_id:
        ultimos_registros = ultimos_registros.filter(servicio__proveedor_id=proveedor_id)
    if establecimiento_id:
        ultimos_registros = ultimos_registros.filter(servicio__establecimiento_id=establecimiento_id)
    if servicio_id:
        ultimos_registros = ultimos_registros.filter(servicio_id=servicio_id)
    
    # Ordenar y limitar
    ultimos_registros = ultimos_registros.order_by('-id_registro')[:20]
    
    # Obtener listas para los selectores de filtro
    proveedores = Proveedor.objects.all().order_by('nombre')
    establecimientos = Establecimientos.objects.all().order_by('nombre')
    servicios = Servicios.objects.select_related('proveedor', 'establecimiento').order_by('numero_servicio')
    
    context = {
        'total_registros': total_registros,
        'total_servicios': total_servicios,
        'total_establecimientos': total_establecimientos,
        'total_proveedores': total_proveedores,
        'cantidad_mes': cantidad_mes,
        'ultimos_registros': ultimos_registros,
        'proveedores': proveedores,
        'establecimientos': establecimientos,
        'servicios': servicios,
        'proveedor_seleccionado': proveedor_id,
        'establecimiento_seleccionado': establecimiento_id,
        'servicio_seleccionado': servicio_id,
    }
    
    return render(request, 'docs/dashboard.html', context)

#vista para descargar planilla masiva de servicios
@login_required
def descargar_plantilla_servicios(request):
    # Crear un DataFrame con las columnas necesarias
    df = pd.DataFrame(columns=[
        'numero_servicio',
        'rbd',
        'rut_proveedor',
        'tipo_recibo'
    ])
    
    # Crear un ejemplo de datos
    ejemplo = pd.DataFrame({
        'numero_servicio': ['SV-001'],
        'rbd': ['12345'],
        'rut_proveedor': ['12345678-9'],
        'tipo_recibo': ['Factura']
    })
    
    # Concatenar el ejemplo con el DataFrame vacío
    df = pd.concat([df, ejemplo], ignore_index=True)
    
    # Crear un archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Plantilla')
        
        # Obtener la hoja de trabajo
        worksheet = writer.sheets['Plantilla']
        
        # Ajustar el ancho de las columnas
        for i, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.column_dimensions[chr(65 + i)].width = max_length + 2
    
    # Preparar la respuesta
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_servicios.xlsx'
    
    return response


#vistas de servicios
@login_required
def listar_servicios(request):
    servicios = Servicios.objects.select_related(
        'establecimiento',
        'proveedor',
        'tipo_recibo'
    ).order_by('establecimiento__nombre', 'numero_servicio')
    return render(request, 'docs/servicios_lista.html', {'servicios': servicios})

@login_required
def crear_servicio(request):
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Servicio creado exitosamente')
            return redirect('docs:listar_servicios')
    else:
        form = ServicioForm()
    return render(request, 'docs/servicio_form.html', {'form': form})

@login_required
def editar_servicio(request, servicio_id):
    servicio = get_object_or_404(Servicios, id_serv=servicio_id)
    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, 'Servicio actualizado exitosamente')
            return redirect('docs:listar_servicios')
    else:
        form = ServicioForm(instance=servicio)
    return render(request, 'docs/servicio_editar.html', {'form': form})

@login_required
def eliminar_servicio(request, servicio_id):
    servicio = get_object_or_404(Servicios, id_serv=servicio_id)
    servicio.delete()
    messages.success(request, 'Servicio eliminado exitosamente')
    return redirect('docs:listar_servicios')

@login_required
def ver_servicio(request, servicio_id):
    servicio = Servicios.objects.get(id_serv=servicio_id)
    return render(request, 'docs/servicio_detalle.html', {'servicio': servicio})


#vistas de establecimientos
@login_required
def listar_establecimientos(request):
    establecimientos = Establecimientos.objects.all().order_by('nombre')
    return render(request, 'docs/establecimientos_lista.html', {'establecimientos': establecimientos})

@login_required
def toggle_establecimiento_activo(request, establecimiento_id):
    """
    Vista para cambiar el estado activo de un establecimiento
    """
    if request.method == 'POST':
        try:
            establecimiento = get_object_or_404(Establecimientos, id_est=establecimiento_id)
            establecimiento.activo = not establecimiento.activo
            establecimiento.save()
            
            return JsonResponse({
                'success': True,
                'activo': establecimiento.activo,
                'message': f'Establecimiento {"activado" if establecimiento.activo else "desactivado"} exitosamente'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al cambiar el estado: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

@login_required
def ver_establecimiento(request, id_est):
    establecimiento = get_object_or_404(Establecimientos, id_est=id_est)
    servicios = Servicios.objects.filter(establecimiento=establecimiento)
    
    context = {
        'establecimiento': establecimiento,
        'servicios': servicios,
    }
    return render(request, 'docs/establecimiento_detalle.html', context)

@login_required
def editar_establecimiento(request, establecimiento_id):
    establecimiento = get_object_or_404(Establecimientos, id_est=establecimiento_id)
    if request.method == 'POST':
        form = EstablecimientoForm(request.POST, instance=establecimiento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Establecimiento actualizado exitosamente')
            return redirect('docs:listar_establecimientos')
    else:
        form = EstablecimientoForm(instance=establecimiento)
    return render(request, 'docs/establecimiento_editar.html', {'form': form})

@login_required
def crear_establecimiento(request):
    if request.method == 'POST':
        form = EstablecimientoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Establecimiento creado exitosamente')
            return redirect('docs:listar_establecimientos')
    else:
        form = EstablecimientoForm()
    return render(request, 'docs/establecimiento_form.html', {'form': form})

@login_required
def eliminar_establecimiento(request, id_est):
    establecimiento = get_object_or_404(Establecimientos, id_est=id_est)
    establecimiento.delete()
    messages.success(request, 'Establecimiento eliminado exitosamente')
    return redirect('docs:listar_establecimientos')  


#vista de proveedores
@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor creado exitosamente')
            return redirect('docs:listar_proveedores')   
    else:
        form = ProveedorForm()
    return render(request, 'docs/proveedor_form.html', {'form': form})

@login_required
def listar_proveedores(request):
    proveedores = Proveedor.objects.all()
    return render(request, 'docs/proveedores_lista.html', {'proveedores': proveedores})

@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado exitosamente')
            return redirect('docs:listar_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'docs/proveedor_editar.html', {'form': form})        


#vistas para registros
@login_required
def importar_registros(request):
    registros_importados = []
    errores = []
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if archivo:
            try:
                start_time = time.time()
                df = pd.read_excel(archivo)
                total = len(df)
                
                messages.info(request, f'Validando {total} registros...')
                
                # Primera pasada: validación completa
                registros_validados = []
                for index, row in df.iterrows():
                    try:
                        # Validar que todos los campos requeridos existan
                        campos_requeridos = ['numero_servicio', 'numero_recibo', 'fecha_envio_pago', 'fecha_vencimiento', 'monto']
                        if not all(k in row for k in campos_requeridos):
                            campos_faltantes = [campo for campo in campos_requeridos if campo not in row]
                            raise ValueError(f"Faltan campos requeridos: {', '.join(campos_faltantes)}")
                            
                        # Validar servicio
                        servicios = Servicios.objects.filter(numero_servicio=row['numero_servicio'])
                        if servicios.count() > 1:
                            raise ValueError(f"Hay {servicios.count()} servicios con el mismo número")
                        elif servicios.count() == 0:
                            raise ValueError(f"No existe un servicio con número {row['numero_servicio']}")
                            
                        # Validar fechas
                        try:
                            # Obtener los valores originales de las fechas
                            fecha_envio_valor = row['fecha_envio_pago']
                            fecha_emision_valor = row.get('fecha_emision')  # Usar get para permitir que no exista
                            fecha_vencimiento_valor = row['fecha_vencimiento']
                            
                            # Verificar si las fechas requeridas están vacías
                            if pd.isna(fecha_envio_valor):
                                raise ValueError("La fecha de envío a pago está vacía. Por favor, ingrese una fecha válida en formato dd-mm-yyyy")
                            if pd.isna(fecha_vencimiento_valor):
                                raise ValueError("La fecha de vencimiento está vacía. Por favor, ingrese una fecha válida en formato dd-mm-yyyy")
                            
                            # Función auxiliar para convertir fechas
                            def parse_fecha(fecha_valor):
                                if isinstance(fecha_valor, datetime):
                                    return fecha_valor.date()
                                elif isinstance(fecha_valor, str):
                                    try:
                                        return datetime.strptime(fecha_valor.strip(), '%d-%m-%Y').date()
                                    except:
                                        # Si falla, intentar con el formato que viene de Excel
                                        return pd.to_datetime(fecha_valor).date()
                                else:
                                    # Si es un timestamp de pandas
                                    return pd.to_datetime(fecha_valor).date()
                            
                            # Convertir fechas
                            fecha_envio = parse_fecha(fecha_envio_valor)
                            fecha_vencimiento = parse_fecha(fecha_vencimiento_valor)
                            
                            # Manejar fecha de emisión si existe
                            fecha_emision = None
                            if not pd.isna(fecha_emision_valor):
                                try:
                                    fecha_emision = parse_fecha(fecha_emision_valor)
                                    
                                    # Validar que la fecha de emisión sea anterior a la fecha de vencimiento
                                    if fecha_emision > fecha_vencimiento:
                                        raise ValueError(f"La fecha de emisión ({fecha_emision}) no puede ser posterior a la fecha de vencimiento ({fecha_vencimiento})")
                                    
                                    # Validar que la fecha de envío sea posterior a la fecha de emisión
                                    if fecha_envio < fecha_emision:
                                        raise ValueError(f"La fecha de envío a pago ({fecha_envio}) no puede ser anterior a la fecha de emisión ({fecha_emision})")
                                except Exception as e:
                                    raise ValueError(f"Error en formato de fecha de emisión: {str(e)}")
                            
                        except ValueError as e:
                            raise ValueError(str(e))
                            
                        # Validar monto
                        try:
                            # Intentar convertir el monto a string y limpiarlo
                            monto_str = str(row['monto']).strip()
                            # Eliminar cualquier carácter que no sea número
                            monto_str = ''.join(c for c in monto_str if c.isdigit())
                            # Convertir a int
                            monto = int(monto_str)
                            if monto <= 0:
                                raise ValueError("El monto debe ser mayor que 0")
                        except (ValueError, TypeError):
                            raise ValueError("El monto debe ser un número entero válido")
                            
                        # Validar número de recibo
                        if not row['numero_recibo'] or len(str(row['numero_recibo']).strip()) == 0:
                            raise ValueError("El número de recibo no puede estar vacío")
                            
                        # Validar interés si existe
                        interes = 0  # Valor por defecto
                        if 'interes' in row and pd.notna(row['interes']):
                            try:
                                interes = int(row['interes'])
                                if interes < 0:
                                    raise ValueError("El interés no puede ser negativo")
                            except ValueError:
                                raise ValueError("El interés debe ser un número entero válido")
                        
                        # Si todo está válido, guardar el registro para importación
                        registros_validados.append({
                            'row': row,
                            'servicio': servicios.first(),
                            'fecha_envio': fecha_envio,
                            'fecha_emision': fecha_emision,  # Puede ser None
                            'fecha_vencimiento': fecha_vencimiento,
                            'interes': interes  # Agregamos el interés al diccionario
                        })
                        
                    except Exception as e:
                        errores.append({
                            'fila': index + 2,  # +2 porque Excel empieza en 1 y la primera fila son los encabezados
                            'mensaje': str(e)
                        })
                
                # Si hay errores, mostrarlos y no proceder con la importación
                if errores:
                    messages.error(request, f'Se encontraron {len(errores)} errores en el archivo. La importación fue cancelada.')
                    return render(request, 'docs/importar_registros.html', {
                        'errores': errores
                    })
                
                # Segunda pasada: importación (solo si no hay errores)
                with transaction.atomic():
                    for registro in registros_validados:
                        row = registro['row']
                        registro_obj = RegistroServicio.objects.create(
                            servicio=registro['servicio'],
                            numero_recibo=row['numero_recibo'],
                            fecha_envio_pago=registro['fecha_envio'],
                            fecha_emision=registro['fecha_emision'],  # Puede ser None
                            fecha_vencimiento=registro['fecha_vencimiento'],
                            monto=int(row['monto']),
                            interes=registro['interes']  # Usamos el interés del diccionario
                        )
                        registros_importados.append(registro_obj)
                
                end_time = time.time()
                tiempo_total = round(end_time - start_time, 2)
                
                messages.success(request, f'''
                    <strong>Importación completada exitosamente:</strong><br>
                    Total procesados: {total}<br>
                    Registros importados: {len(registros_importados)}<br>
                    Tiempo de procesamiento: {tiempo_total} segundos
                ''')
                
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
        else:
            messages.error(request, 'No se ha seleccionado ningún archivo')
        
        return render(request, 'docs/importar_registros.html', {
            'registros_importados': registros_importados,
            'errores': errores
        })
    
    return render(request, 'docs/importar_registros.html')

@login_required
def descargar_plantilla_registros(request):
    """Vista para descargar la plantilla de importación de registros"""
    # Crear un DataFrame con las columnas necesarias
    columns = [
        'numero_servicio',
        'numero_recibo',
        'fecha_envio_pago',
        'fecha_emision',
        'fecha_vencimiento',
        'monto',
        'interes'
    ]
    df = pd.DataFrame(columns=columns)
    
    # Agregar una fila de ejemplo
    ejemplo = pd.DataFrame({
        'numero_servicio': ['SV-001'],
        'numero_recibo': ['F-001'],
        'fecha_envio_pago': ['01-01-2024'],
        'fecha_emision': ['01-01-2024'],
        'fecha_vencimiento': ['31-01-2024'],
        'monto': [100000],
        'interes': [0]
    })
    
    # Concatenar el ejemplo con el DataFrame vacío
    df = pd.concat([df, ejemplo], ignore_index=True)
    
    # Crear el buffer en memoria para el archivo Excel
    buffer = io.BytesIO()
    
    # Guardar el DataFrame como Excel en el buffer
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Plantilla')
        
        # Obtener la hoja de trabajo
        worksheet = writer.sheets['Plantilla']
        
        # Ajustar el ancho de las columnas
        for i, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.column_dimensions[chr(65 + i)].width = max_length + 2
    
    # Preparar la respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_registros.xlsx'
    
    return response

@login_required
def listar_registros(request):
    establecimientos = Establecimientos.objects.filter(activo=True).order_by('nombre')
    establecimiento_id = request.GET.get('establecimiento')
    
    # Si se intenta acceder con el valor 'jardines', redirigir a la nueva funcionalidad
    if establecimiento_id == 'jardines':
        return redirect('docs:descargar_pdfs_jardines')
    
    registros = []
    if establecimiento_id:
        try:
            # Validar que el establecimiento_id sea un número válido
            establecimiento_id = int(establecimiento_id)
            # Obtener registros del establecimiento seleccionado
            registros = RegistroServicio.objects.filter(
                servicio__establecimiento_id=establecimiento_id
            ).select_related(
                'servicio', 
                'servicio__establecimiento',
                'servicio__proveedor',
                'servicio__proveedor__tipo_proveedor'
            ).order_by('-fecha_emision')
        except (ValueError, TypeError):
            # Si el establecimiento_id no es un número válido, limpiar la selección
            establecimiento_id = None
    
    return render(request, 'docs/registros_lista.html', {
        'establecimientos': establecimientos,
        'registros': registros,
        'establecimiento_seleccionado': establecimiento_id
    })

@login_required
def descargar_registro_pdf(request, registro_id):
    # Configurar locale para fechas en español
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except:
        # Si falla, usaremos una lista manual de meses
        MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']

    # Obtener el registro y sus relaciones
    registro = get_object_or_404(RegistroServicio.objects.select_related(
        'servicio',
        'servicio__establecimiento',
        'servicio__proveedor',
        'servicio__tipo_recibo'
    ), id_registro=registro_id)

    # Crear el buffer de memoria para el PDF
    buffer = io.BytesIO()
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,    # Reducido de 50 a 30
        leftMargin=50,     # Reducido de 50 a 30
        topMargin=50,      # Reducido de 50 a 30
        bottomMargin=50    # Reducido de 50 a 30
    )

    # Contenedor para los elementos del PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='MainTitle',
        parent=styles['Heading1'],
        alignment=TA_CENTER,
        fontSize=12,     # Reducido de 14 a 12
        spaceAfter=15,   # Reducido de 20 a 15
        spaceBefore=15   # Reducido de 20 a 15
    ))
    styles.add(ParagraphStyle(
        name='SignatureTitle',
        parent=styles['Heading1'],
        alignment=TA_CENTER,
        fontSize=10,     # Reducido de 12 a 10
        spaceAfter=4,    # Reducido de 5 a 4
        spaceBefore=4    # Reducido de 5 a 4
    ))
    styles.add(ParagraphStyle(
        name='NormalText',
        parent=styles['Normal'],
        fontSize=10,     # Reducido de 12 a 10
        leading=14,      # Reducido de 16 a 14
        spaceBefore=10,  # Reducido de 12 a 10
        spaceAfter=10    # Reducido de 12 a 10
    ))

    # Obtener rutas de los logos
    BASE_DIR = Path(__file__).resolve().parent
    logo_izq = os.path.join(BASE_DIR, 'static', 'img', 'logo_slep.png')
    logo_der = os.path.join(BASE_DIR, 'static', 'img', 'Logo.png')

    try:
        # Crear tabla de encabezado con logos
        logo_width_izq = 0.85*inch
        logo_height_izq = 0.85*inch
        logo_width_der = 1.7*inch
        logo_height_der = 1*inch

        # Intentar cargar ambos logos
        logo_izq_img = Image(logo_izq, width=logo_width_izq, height=logo_height_izq)
        logo_der_img = Image(logo_der, width=logo_width_der, height=logo_height_der)

        header_data = [[
            logo_izq_img,
            Paragraph("", styles['Normal']),
            logo_der_img
        ]]

        header_table = Table(header_data, colWidths=[2*inch, 3*inch, 2*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (2,0), (2,0), 'RIGHT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),    # Alineación vertical superior para todos
            ('TOPPADDING', (0,0), (-1,-1), 0),    # Sin padding superior
            ('BOTTOMPADDING', (0,0), (-1,-1), 0), # Sin padding inferior
        ]))

        elements.append(header_table)
        elements.append(Spacer(1, 40))  # Aumentado espacio después de los logos

        # Título
        elements.append(Paragraph("RECEPCIÓN CONFORME", styles['MainTitle']))

        # Texto introductorio
        fecha_actual = datetime.now()
        try:
            mes = fecha_actual.strftime('%B').lower()  # Intentar obtener el mes en español
        except:
            mes = MESES[fecha_actual.month - 1]  # Usar la lista manual si falla
        
        texto_intro = f"""En Iquique, a {fecha_actual.day} de {mes} de {fecha_actual.year} en el establecimiento {registro.servicio.establecimiento.nombre}, se procede a dar recepción conforme a las boletas de {registro.servicio.proveedor.nombre}, se adjunta listado."""
        elements.append(Paragraph(texto_intro, styles['NormalText']))
        elements.append(Spacer(1, 20))

        # Definir colores personalizados
        azul_oscuro = HexColor('#1F4970')  # Azul corporativo
        gris_claro = HexColor('#F5F5F5')   # Gris para fondo
        gris_lineas = HexColor('#CCCCCC')  # Gris para líneas

        # Función para formatear números con punto como separador de miles
        def format_number(number):
            return f'${number:,.0f}'.replace(',', '.')

        # Preparar los datos con Paragraph para permitir wrapping
        data = [
            ['N° Cliente', 'Establecimiento', 'Factura', 'Monto JUNJI', 'Monto Final'],
            [
                registro.servicio.numero_servicio,
                Paragraph(registro.servicio.establecimiento.nombre, 
                         ParagraphStyle(
                             'EstablishmentStyle',
                             parent=styles['Normal'],
                             fontSize=9,
                             leading=11,
                             wordWrap='LTR',  # Left to Right wrapping
                             splitLongWords=False,  # No partir palabras
                             spaceShrinkage=0.05,  # Permitir menos espacio entre palabras
                         )),
                registro.numero_recibo,
                format_number(registro.monto - (registro.interes or 0)),
                format_number(registro.monto)
            ]
        ]

        # Calcular anchos de columna para que coincidan con los márgenes del documento
        available_width = letter[0] - 100  # Ancho total menos márgenes (50 por lado)
        col_widths = [
            available_width * 0.12,  # N° Cliente
            available_width * 0.41,  # Establecimiento
            available_width * 0.17,  # Factura
            available_width * 0.15,  # Monto JUNJI
            available_width * 0.15   # Monto Final
        ]

        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Alineación
            ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Centrar encabezados
            ('ALIGN', (0,1), (2,1), 'LEFT'),     # Alinear texto a la izquierda
            ('ALIGN', (3,1), (-1,-1), 'RIGHT'),  # Alinear montos a la derecha
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),# Alineación vertical central
            
            # Fuentes y tamaños
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),  # Encabezados en negrita
            ('FONTNAME', (0,1), (-1,1), 'Helvetica'),       # Contenido normal
            ('FONTSIZE', (0,0), (-1,-1), 9),                # Tamaño base de fuente
            
            # Colores y bordes
            ('BACKGROUND', (0,0), (-1,0), azul_oscuro),     # Fondo azul en encabezados
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),     # Texto blanco en encabezados
            ('BACKGROUND', (0,1), (-1,-1), gris_claro),     # Fondo gris claro en contenido
            ('GRID', (0,0), (-1,-1), 0.5, gris_lineas),    # Líneas de grid suaves
            ('BOX', (0,0), (-1,-1), 1, azul_oscuro),       # Borde exterior más grueso
            
            # Espaciado diferenciado para cabecera y contenido
            ('TOPPADDING', (0,0), (-1,0), 6),     # Padding superior reducido para cabecera
            ('BOTTOMPADDING', (0,0), (-1,0), 6),  # Padding inferior reducido para cabecera
            ('TOPPADDING', (0,1), (-1,-1), 12),   # Padding superior normal para contenido
            ('BOTTOMPADDING', (0,1), (-1,-1), 12), # Padding inferior normal para contenido
            ('LEFTPADDING', (0,0), (-1,-1), 10),   # Padding izquierdo consistente
            ('RIGHTPADDING', (0,0), (-1,-1), 10),  # Padding derecho consistente
            
            # Wrapping
            ('WORDWRAP', (1,1), (1,1), True),  # Asegurar wrapping en nombre del establecimiento
        ]))

        # Agregar la tabla al documento con espaciado adicional
        elements.append(table)
        elements.append(Spacer(1, 180))  # Aumentado de 120 a 180 puntos para bajar más la firma
        
        # Línea de firma más elegante
        signature_width = 200  # Ancho de la línea de firma
        signature_line = Table([['']], colWidths=[signature_width])
        signature_line.setStyle(TableStyle([
            ('LINEABOVE', (0,0), (-1,-1), 1, azul_oscuro),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
        ]))
        elements.append(signature_line)
        elements.append(Spacer(1, 5))
        
        # Texto "RECIBE CONFORME" más pequeño y más cerca de la línea
        elements.append(Paragraph("RECIBE CONFORME", styles['SignatureTitle']))
        elements.append(Spacer(1, 60))  # Aumentado de 40 a 60 puntos el espacio final

        # Construir el PDF
        doc.build(elements)
        
        # Preparar la respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        # Nombre del archivo: {nombre establecimiento}_{numero factura o recibo}
        nombre_archivo = f"{registro.servicio.establecimiento.nombre}_{registro.numero_recibo}.pdf"
        # Limpiar caracteres no válidos para nombres de archivo
        nombre_archivo = "".join(c for c in nombre_archivo if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
        nombre_archivo = nombre_archivo.replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        return response

    except Exception as e:
        messages.error(request, f'Error al generar el PDF: {str(e)}')
        return render(request, 'docs/error.html')

@login_required
def generar_enlace_correo(request, registro_id):
    registro = get_object_or_404(RegistroServicio.objects.select_related(
        'servicio',
        'servicio__establecimiento',
        'servicio__establecimiento__director',
        'servicio__proveedor'  # Cambiamos a servicio__proveedor
    ), id_registro=registro_id)

    try:
        # Obtener los correos
        establecimiento = registro.servicio.establecimiento
        director = establecimiento.director
        destinatarios = []
        
        if establecimiento.email:
            destinatarios.append(establecimiento.email)
        if director and director.email:
            destinatarios.append(director.email)
        
        # Crear el asunto del correo usando el acrónimo del proveedor
        asunto = "{} - Envío de Boletas y Documentos de Recepción para Firma".format(registro.servicio.proveedor.acronimo)

        # Crear el cuerpo del correo
        cuerpo = """Estimado(a):

Esperando que se encuentre bien. Adjunto las boletas de los servicios básicos junto con los documentos de recepción conforme correspondientes a cada tipo de servicio. Estos deben ser firmados por el director.
Una vez que recibamos los documentos firmados, procederemos con el pago correspondiente.
Favor recordar que estos documentos se deben enviar de forma digital.

Agradezco de antemano su pronta atención a este asunto.

Saludos cordiales."""
        copia = ('ssgg@slepiquique.cl','katherine.cerda@slepiquique.cl')

        # Crear el enlace mailto
        mailto_link = (
            f"mailto:{','.join(destinatarios)}"
            f"?cc={','.join(copia)}"
            f"?subject={quote(asunto)}"
            f"&body={quote(cuerpo)}"
        )
        
        return HttpResponse(mailto_link)
        
    except Exception as e:
        messages.error(request, f'Error al generar el enlace de correo: {str(e)}')
        return HttpResponse('Error al generar el enlace de correo', status=500)

@login_required
def crear_registro(request):
    establecimiento_id = request.GET.get('establecimiento')
    
    if request.method == 'POST':
        form = RegistroServicioForm(request.POST, establecimiento_id=establecimiento_id)
        if form.is_valid():
            registro = form.save()
            messages.success(request, 'Registro creado exitosamente')
            return redirect('docs:listar_registros')
    else:
        form = RegistroServicioForm(establecimiento_id=establecimiento_id)
    
    # Obtener establecimientos para el select
    establecimientos = Establecimientos.objects.all().order_by('nombre')
    
    return render(request, 'docs/registro_form.html', {
        'form': form,
        'establecimientos': establecimientos,
        'establecimiento_seleccionado': establecimiento_id
    })

@login_required
def descargar_pdfs_jardines(request):
    """
    Vista para descargar PDFs del último registro de cada jardín
    """
    # Filtrar solo jardines activos (establecimientos que contengan palabras clave de jardines y estén activos)
    jardines = Establecimientos.objects.filter(
        models.Q(nombre__icontains='jardín') |
        models.Q(nombre__icontains='jardin') |
        models.Q(nombre__icontains='junji') |
        models.Q(nombre__icontains='infantil') |
        models.Q(nombre__icontains='niños') |
        models.Q(nombre__icontains='niñas') |
        models.Q(nombre__icontains='preescolar'),
        activo=True
    ).order_by('nombre')
    
    proveedor_id = request.GET.get('proveedor')
    
    # Obtener el último registro de cada jardín usando Python
    from django.db.models import Max
    
    # Obtener todos los registros de jardines ordenados por fecha de vencimiento (más reciente primero)
    todos_registros = RegistroServicio.objects.filter(
        servicio__establecimiento__in=jardines
    ).select_related(
        'servicio', 
        'servicio__establecimiento',
        'servicio__proveedor',
        'servicio__proveedor__tipo_proveedor'
    ).order_by('servicio__establecimiento', '-fecha_vencimiento')
    
    # Agrupar por establecimiento Y proveedor, tomar el primero (más reciente) de cada grupo
    ultimos_registros = []
    combinaciones_vistas = set()
    
    for registro in todos_registros:
        # Crear una clave única para establecimiento + proveedor
        clave = f"{registro.servicio.establecimiento.id_est}_{registro.servicio.proveedor.id_prov}"
        if clave not in combinaciones_vistas:
            ultimos_registros.append(registro)
            combinaciones_vistas.add(clave)
    
    # Aplicar filtro por proveedor si existe
    if proveedor_id:
        ultimos_registros = [r for r in ultimos_registros if r.servicio.proveedor.id_prov == int(proveedor_id)]
    
    # Obtener TODOS los proveedores que tienen registros en jardines (no solo los filtrados)
    todos_proveedores_ids = RegistroServicio.objects.filter(
        servicio__establecimiento__in=jardines
    ).values_list('servicio__proveedor__id_prov', flat=True).distinct()

    proveedores = Proveedor.objects.filter(id_prov__in=todos_proveedores_ids).order_by('nombre')


    return render(request, 'docs/descargar_pdfs_jardines.html', {
        'jardines': jardines,
        'registros': ultimos_registros,
        'proveedores': proveedores,
        'proveedor_seleccionado': proveedor_id
    })

def generar_pdf_registro(registro):
    """
    Función auxiliar para generar PDF de un registro individual
    Reutiliza la misma lógica que descargar_registro_pdf
    """
    import io
    import locale
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.utils import ImageReader
    from django.conf import settings
    import os
    
    # Configurar locale para fechas en español
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except:
        MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']

    # Crear el buffer de memoria para el PDF
    buffer = io.BytesIO()
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )

    # Contenedor para los elementos del PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='MainTitle',
        parent=styles['Heading1'],
        alignment=TA_CENTER,
        fontSize=12,
        spaceAfter=15,
        spaceBefore=15
    ))
    styles.add(ParagraphStyle(
        name='SignatureTitle',
        parent=styles['Heading1'],
        alignment=TA_CENTER,
        fontSize=10,
        spaceAfter=4,
        spaceBefore=4
    ))

    # Función para formatear números
    def format_number(value):
        if value is None:
            return "0"
        return f"${value:,.0f}".replace(',', '.')

    # Cargar y redimensionar logos
    try:
        logo_izq_path = os.path.join(settings.STATIC_ROOT, 'img', 'logo_slep.png')
        logo_der_path = os.path.join(settings.STATIC_ROOT, 'img', 'Logo.png')
        
        if os.path.exists(logo_izq_path):
            logo_izq = ImageReader(logo_izq_path)
            logo_width_izq = 1.5*inch
            logo_height_izq = 0.8*inch
            logo_izq_img = Image(logo_izq, width=logo_width_izq, height=logo_height_izq)
        else:
            logo_izq_img = Paragraph("", styles['Normal'])
            
        if os.path.exists(logo_der_path):
            logo_der = ImageReader(logo_der_path)
            logo_width_der = 1.5*inch
            logo_height_der = 0.8*inch
            logo_der_img = Image(logo_der, width=logo_width_der, height=logo_height_der)
        else:
            logo_der_img = Paragraph("", styles['Normal'])
    except:
        logo_izq_img = Paragraph("", styles['Normal'])
        logo_der_img = Paragraph("", styles['Normal'])

    # Header con logos
    header_data = [[
        logo_izq_img,
        Paragraph("", styles['Normal']),
        logo_der_img
    ]]

    header_table = Table(header_data, colWidths=[2*inch, 3*inch, 2*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('ALIGN', (2,0), (2,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))

    # Título principal
    elements.append(Paragraph("RECEPCIÓN CONFORME", styles['MainTitle']))
    elements.append(Spacer(1, 20))

    # Preparar los datos con Paragraph para permitir wrapping
    data = [
        ['N° Cliente', 'Establecimiento', 'Factura', 'Monto JUNJI', 'Monto Final'],
        [
            registro.servicio.numero_servicio,
            Paragraph(registro.servicio.establecimiento.nombre, 
                     ParagraphStyle(
                         'EstablishmentStyle',
                         parent=styles['Normal'],
                         fontSize=9,
                         leading=11,
                         wordWrap='LTR',
                         splitLongWords=False,
                         spaceShrinkage=0.05,
                     )),
            registro.numero_recibo,
            format_number(registro.monto - (registro.interes or 0)),
            format_number(registro.monto)
        ]
    ]

    # Calcular anchos de columna
    available_width = letter[0] - 100
    col_widths = [
        available_width * 0.15,  # N° Cliente
        available_width * 0.35,  # Establecimiento
        available_width * 0.15,  # Factura
        available_width * 0.175, # Monto JUNJI
        available_width * 0.175  # Monto Final
    ]

    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 30))

    # Información adicional
    elements.append(Paragraph(f"<b>Fecha de Emisión:</b> {registro.fecha_emision.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Paragraph(f"<b>Fecha de Vencimiento:</b> {registro.fecha_vencimiento.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Paragraph(f"<b>Proveedor:</b> {registro.servicio.proveedor.nombre}", styles['Normal']))
    elements.append(Paragraph(f"<b>Tipo de Servicio:</b> {registro.servicio.proveedor.tipo_proveedor.nombre}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Sección de firmas
    elements.append(Paragraph("FIRMAS", styles['SignatureTitle']))
    elements.append(Spacer(1, 10))

    # Tabla de firmas
    signature_data = [
        ['Representante Legal', 'Contador', 'Director'],
        ['', '', ''],
        ['', '', ''],
        ['', '', ''],
    ]

    signature_table = Table(signature_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, 1), (0, 3), 1, colors.black),
        ('LINEBELOW', (1, 1), (1, 3), 1, colors.black),
        ('LINEBELOW', (2, 1), (2, 3), 1, colors.black),
    ]))

    elements.append(signature_table)

    # Construir PDF
    doc.build(elements)
    
    # Obtener bytes del PDF
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes

@login_required
def descargar_zip_jardines(request):
    """
    Vista para descargar un ZIP con todos los PDFs de jardines
    """
    import zipfile
    import tempfile
    import os
    import io
    import locale
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Filtrar solo jardines activos
    jardines = Establecimientos.objects.filter(
        models.Q(nombre__icontains='jardín') |
        models.Q(nombre__icontains='jardin') |
        models.Q(nombre__icontains='junji') |
        models.Q(nombre__icontains='infantil') |
        models.Q(nombre__icontains='niños') |
        models.Q(nombre__icontains='niñas') |
        models.Q(nombre__icontains='preescolar'),
        activo=True
    ).order_by('nombre')
    
    proveedor_id = request.GET.get('proveedor')
    
    # Obtener todos los registros de jardines ordenados por fecha de vencimiento
    todos_registros = RegistroServicio.objects.filter(
        servicio__establecimiento__in=jardines
    ).select_related(
        'servicio', 
        'servicio__establecimiento',
        'servicio__proveedor',
        'servicio__proveedor__tipo_proveedor'
    ).order_by('servicio__establecimiento', '-fecha_vencimiento')
    
    # Agrupar por establecimiento Y proveedor, tomar el primero (más reciente) de cada grupo
    ultimos_registros = []
    combinaciones_vistas = set()
    
    for registro in todos_registros:
        clave = f"{registro.servicio.establecimiento.id_est}_{registro.servicio.proveedor.id_prov}"
        if clave not in combinaciones_vistas:
            ultimos_registros.append(registro)
            combinaciones_vistas.add(clave)
    
    # Aplicar filtro por proveedor si existe
    if proveedor_id:
        ultimos_registros = [r for r in ultimos_registros if r.servicio.proveedor.id_prov == int(proveedor_id)]
    
    if not ultimos_registros:
        return HttpResponse("No hay registros de jardines para descargar.", status=404)
    
    # Configurar locale para fechas en español
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except:
        MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    
    # Crear archivo ZIP temporal
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    
    try:
        with zipfile.ZipFile(temp_file.name, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for i, registro in enumerate(ultimos_registros, 1):
                try:
                    # Usar la función generar_pdf_registro para mantener consistencia
                    pdf_bytes = generar_pdf_registro(registro)
                    
                    # Nombre del archivo: {nombre establecimiento}_{numero factura o recibo}
                    establecimiento = registro.servicio.establecimiento
                    nombre_archivo = f"{establecimiento.nombre}_{registro.numero_recibo}.pdf"
                    # Limpiar caracteres no válidos para nombres de archivo
                    nombre_archivo = "".join(c for c in nombre_archivo if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
                    nombre_archivo = nombre_archivo.replace(' ', '_')
                    
                    # Agregar al ZIP
                    zip_file.writestr(nombre_archivo, pdf_bytes)
                    
                except Exception as e:
                    print(f"Error generando PDF para registro {registro.id_registro}: {str(e)}")
                    continue
        
        # Leer el archivo ZIP
        with open(temp_file.name, 'rb') as f:
            zip_data = f.read()
        
        # Crear respuesta
        response = HttpResponse(zip_data, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="pdfs_jardines.zip"'
        response['Content-Length'] = len(zip_data)
        
        return response
        
    finally:
        # Limpiar archivo temporal
        try:
            os.unlink(temp_file.name)
        except:
            pass

@login_required
def perfil_usuario(request):
    if request.method == 'POST':
        form = PerfilUsuarioForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('docs:perfil_usuario')
    else:
        form = PerfilUsuarioForm(instance=request.user)
    
    return render(request, 'docs/perfil_usuario.html', {'form': form})


class CambiarPasswordView(PasswordChangeView, LoginRequiredMixin):
    template_name = 'docs/cambiar_password.html'
    form_class = PasswordChangeForm
    success_url = reverse_lazy('docs:perfil_usuario')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Cambiar Contraseña'
        return context

def descargarMasivoRegistros(request):
    # CORRECTO: Usar .all() para obtener todos los registros
    registros = RegistroServicio.objects.select_related(
        'servicio',
        'servicio__establecimiento',
        'servicio__proveedor',
        'servicio__tipo_recibo',
    ).all()

    columns = [
        'Proveedor',
        'Rut proveedor',
        'Establecimiento',
        'RBD',
        'Numero de Servicio',
        'Tipo de documento',
        'Numero de documento',
        'Fecha de emisión',
        'Fecha de vencimiento',
        'Monto',
        'Interés'
    ]

    data = []
    for registro in registros:
       
        data.append({
            'Proveedor': registro.servicio.proveedor.nombre, 
            'Rut proveedor': registro.servicio.proveedor.rut,
            'Establecimiento': registro.servicio.establecimiento.nombre, 
            'RBD': registro.servicio.establecimiento.rbd,
            'Numero de Servicio': registro.servicio.numero_servicio,
            'Tipo de documento': registro.servicio.tipo_recibo.nombre, 
            'Numero de documento': registro.numero_recibo,
            'Fecha de emisión': registro.fecha_emision if registro.fecha_emision else None,
            'Fecha de vencimiento': registro.fecha_vencimiento if registro.fecha_vencimiento else None,
            'Monto': int(registro.monto) if registro.monto is not None else 0,
            'Interés': int(registro.interes) if registro.interes is not None else 0,
        })

    df = pd.DataFrame(data, columns=columns)
    
    # Convertir columnas numéricas a tipo int para asegurar formato correcto
    df['Monto'] = pd.to_numeric(df['Monto'], errors='coerce').fillna(0).astype(int)
    df['Interés'] = pd.to_numeric(df['Interés'], errors='coerce').fillna(0).astype(int)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="registros_servicios.xlsx"'

    with io.BytesIO() as b:
        writer = pd.ExcelWriter(b, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Registros')
        
        # Obtener la hoja de trabajo para aplicar formato
        worksheet = writer.sheets['Registros']
        
        # Aplicar formato de número a las columnas numéricas
        from openpyxl.utils import get_column_letter
        
        # Encontrar las columnas de Monto e Interés
        monto_col = None
        interes_col = None
        for i, col in enumerate(columns):
            if col == 'Monto':
                monto_col = get_column_letter(i + 1)
            elif col == 'Interés':
                interes_col = get_column_letter(i + 1)
        
        # Aplicar formato de número con separador de miles
        if monto_col:
            for row in range(2, len(data) + 2):  # Empezar desde la fila 2 (después del encabezado)
                cell = worksheet[f'{monto_col}{row}']
                cell.number_format = '#,##0'
        
        if interes_col:
            for row in range(2, len(data) + 2):  # Empezar desde la fila 2 (después del encabezado)
                cell = worksheet[f'{interes_col}{row}']
                cell.number_format = '#,##0'
        
        # Encontrar las columnas de fechas
        fecha_emision_col = None
        fecha_vencimiento_col = None
        for i, col in enumerate(columns):
            if col == 'Fecha de emisión':
                fecha_emision_col = get_column_letter(i + 1)
            elif col == 'Fecha de vencimiento':
                fecha_vencimiento_col = get_column_letter(i + 1)
        
        # Aplicar formato de fecha a las columnas de fechas
        if fecha_emision_col:
            for row in range(2, len(data) + 2):  # Empezar desde la fila 2 (después del encabezado)
                cell = worksheet[f'{fecha_emision_col}{row}']
                if cell.value is not None:  # Solo aplicar formato si hay valor
                    cell.number_format = 'dd-mm-yyyy'
        
        if fecha_vencimiento_col:
            for row in range(2, len(data) + 2):  # Empezar desde la fila 2 (después del encabezado)
                cell = worksheet[f'{fecha_vencimiento_col}{row}']
                if cell.value is not None:  # Solo aplicar formato si hay valor
                    cell.number_format = 'dd-mm-yyyy'
        
        writer.close()
        response.write(b.getvalue())

    return response

@login_required
def buscar_registro(request):
    """Vista para buscar registros por número de servicio o número de recibo"""
    registros = []
    query = ''
    tipo_busqueda = 'servicio'  # Por defecto buscar por servicio
    
    if request.GET.get('q'):
        query = request.GET.get('q').strip()
        tipo_busqueda = request.GET.get('tipo', 'servicio')
        
        if query:
            if tipo_busqueda == 'servicio':
                # Buscar por número de servicio
                registros = RegistroServicio.objects.filter(
                    servicio__numero_servicio__icontains=query
                ).select_related(
                    'servicio',
                    'servicio__establecimiento',
                    'servicio__proveedor',
                    'servicio__tipo_recibo'
                ).order_by('-fecha_emision')
            else:
                # Buscar por número de recibo
                registros = RegistroServicio.objects.filter(
                    numero_recibo__icontains=query
                ).select_related(
                    'servicio',
                    'servicio__establecimiento',
                    'servicio__proveedor',
                    'servicio__tipo_recibo'
                ).order_by('-fecha_emision')
    
    context = {
        'registros': registros,
        'query': query,
        'tipo_busqueda': tipo_busqueda
    }
    
    return render(request, 'docs/buscar_registro.html', context)

@login_required
def editar_registro(request, registro_id):
    """Vista para editar un registro de servicio"""
    registro = get_object_or_404(RegistroServicio, id_registro=registro_id)
    
    if request.method == 'POST':
        form = RegistroServicioForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro actualizado exitosamente')
            return redirect('docs:buscar_registro')
    else:
        form = RegistroServicioForm(instance=registro)
    
    context = {
        'form': form,
        'registro': registro
    }
    
    return render(request, 'docs/editar_registro.html', context)

@login_required
def proximos_vencer(request):
    """Vista para mostrar registros próximos a vencer"""
    from datetime import timedelta
    from django.utils import timezone
    
    hoy = timezone.now().date()
    dias_filtro = request.GET.get('dias', '7')
    
    try:
        dias = int(dias_filtro)
    except ValueError:
        dias = 7
    
    fecha_limite = hoy + timedelta(days=dias)
    
    # Registros próximos a vencer
    registros_proximos = RegistroServicio.objects.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_limite
    ).select_related(
        'servicio',
        'servicio__establecimiento',
        'servicio__proveedor',
        'servicio__tipo_recibo'
    ).order_by('fecha_vencimiento')
    
    # Registros ya vencidos
    registros_vencidos = RegistroServicio.objects.filter(
        fecha_vencimiento__lt=hoy
    ).select_related(
        'servicio',
        'servicio__establecimiento',
        'servicio__proveedor',
        'servicio__tipo_recibo'
    ).order_by('fecha_vencimiento')
    
    context = {
        'registros_proximos': registros_proximos,
        'registros_vencidos': registros_vencidos,
        'dias_filtro': dias,
        'hoy': hoy
    }
    
    return render(request, 'docs/proximos_vencer.html', context)

@login_required
def reportes_periodo(request):
    """Vista para generar reportes por período basado en fecha de vencimiento"""
    from django.db.models import Sum, Count
    from datetime import datetime
    
    # Obtener parámetros del formulario
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    proveedor_id = request.GET.get('proveedor')
    establecimiento_id = request.GET.get('establecimiento')
    servicio_id = request.GET.get('servicio')
    
    # Inicializar variables
    registros = None
    total_registros = 0
    
    # Si hay fechas, procesar el reporte
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            # Query base - filtrar por fecha de vencimiento
            registros = RegistroServicio.objects.filter(
                fecha_vencimiento__gte=fecha_inicio_obj,
                fecha_vencimiento__lte=fecha_fin_obj
            ).select_related(
                'servicio',
                'servicio__establecimiento',
                'servicio__proveedor',
                'servicio__tipo_recibo'
            )
            
            # Aplicar filtros adicionales
            if proveedor_id:
                registros = registros.filter(servicio__proveedor_id=proveedor_id)
            if establecimiento_id:
                registros = registros.filter(servicio__establecimiento_id=establecimiento_id)
            if servicio_id:
                registros = registros.filter(servicio_id=servicio_id)
                # Agregar mensaje informativo cuando se filtra por servicio específico
                try:
                    servicio_obj = Servicios.objects.get(id_serv=servicio_id)
                    messages.info(request, f'Mostrando solo registros del servicio: {servicio_obj.numero_servicio}')
                except Servicios.DoesNotExist:
                    pass
            
            # Ordenar por proveedor, establecimiento y fecha de vencimiento
            registros = registros.order_by(
                'servicio__proveedor__nombre',
                'servicio__establecimiento__nombre',
                'fecha_vencimiento'
            )
            
            total_registros = registros.count()
            
        except ValueError:
            messages.error(request, 'Formato de fecha inválido')
    
    # Obtener listas para los filtros
    proveedores = Proveedor.objects.all().order_by('nombre')
    establecimientos = Establecimientos.objects.all().order_by('nombre')
    servicios = Servicios.objects.select_related('proveedor', 'establecimiento').order_by('numero_servicio')
    
    context = {
        'registros': registros,
        'total_registros': total_registros,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'proveedores': proveedores,
        'establecimientos': establecimientos,
        'servicios': servicios,
        'proveedor_seleccionado': proveedor_id,
        'establecimiento_seleccionado': establecimiento_id,
        'servicio_seleccionado': servicio_id,
    }
    
    return render(request, 'docs/reportes_periodo.html', context)

@login_required
def exportar_reporte_periodo(request):
    """Vista para exportar reporte por período a Excel (basado en fecha de vencimiento)"""
    from django.db.models import Sum, Count
    from datetime import datetime
    
    # Obtener parámetros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    proveedor_id = request.GET.get('proveedor')
    establecimiento_id = request.GET.get('establecimiento')
    servicio_id = request.GET.get('servicio')
    
    if not fecha_inicio or not fecha_fin:
        messages.error(request, 'Debe seleccionar un rango de fechas')
        return redirect('docs:reportes_periodo')
    
    try:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Query base - filtrar por fecha de vencimiento
        registros = RegistroServicio.objects.filter(
            fecha_vencimiento__gte=fecha_inicio_obj,
            fecha_vencimiento__lte=fecha_fin_obj
        ).select_related(
            'servicio',
            'servicio__establecimiento',
            'servicio__proveedor',
            'servicio__tipo_recibo'
        )
        
        # Aplicar filtros adicionales
        if proveedor_id:
            registros = registros.filter(servicio__proveedor_id=proveedor_id)
        if establecimiento_id:
            registros = registros.filter(servicio__establecimiento_id=establecimiento_id)
        if servicio_id:
            registros = registros.filter(servicio_id=servicio_id)
        
        # Ordenar por proveedor, establecimiento y fecha de vencimiento
        registros = registros.order_by('servicio__proveedor__nombre', 'servicio__establecimiento__nombre', 'fecha_vencimiento')
        
        # Crear DataFrame con los datos
        data = []
        for registro in registros:
            data.append({
                'Proveedor': registro.servicio.proveedor.nombre,
                'Establecimiento': registro.servicio.establecimiento.nombre,
                'RBD': registro.servicio.establecimiento.rbd,
                'Numero de Servicio': registro.servicio.numero_servicio,
                'Tipo de documento': registro.servicio.tipo_recibo.nombre,
                'Numero de documento': registro.numero_recibo,
                'Fecha de emisión': registro.fecha_emision if registro.fecha_emision else None,
                'Fecha de vencimiento': registro.fecha_vencimiento,
                'Monto': int(registro.monto) if registro.monto is not None else 0,
                'Interés': int(registro.interes) if registro.interes is not None else 0,
                'Total': (int(registro.monto) if registro.monto else 0) + (int(registro.interes) if registro.interes else 0),
            })
        
        df = pd.DataFrame(data)
        
        # Crear buffer para el archivo Excel
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Escribir datos principales
            df.to_excel(writer, index=False, sheet_name='Registros', startrow=4)
            
            # Obtener el workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Registros']
            
            # Importar estilos de openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Agregar título
            worksheet['A1'] = 'REPORTE DE REGISTROS POR PERÍODO'
            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet['A1'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('A1:K1')
            
            # Agregar información del período
            worksheet['A2'] = f'Período: {fecha_inicio_obj.strftime("%d/%m/%Y")} al {fecha_fin_obj.strftime("%d/%m/%Y")}'
            worksheet['A2'].font = Font(size=12)
            worksheet.merge_cells('A2:K2')
            
            # Calcular totales
            total_monto = df['Monto'].sum()
            total_interes = df['Interés'].sum()
            total_general = df['Total'].sum()
            
            # Agregar línea de totales
            worksheet['A3'] = f'Total de registros: {len(df)} | Monto: ${total_monto:,.0f} | Intereses: ${total_interes:,.0f} | Total General: ${total_general:,.0f}'
            worksheet['A3'].font = Font(size=11, bold=True)
            worksheet.merge_cells('A3:K3')
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col in range(1, 12):
                cell = worksheet.cell(row=5, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 25, 'B': 35, 'C': 12, 'D': 18, 'E': 18,
                'F': 18, 'G': 15, 'H': 15, 'I': 12, 'J': 12, 'K': 12
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formatear fechas
            for row in range(6, len(df) + 6):
                # Fecha de emisión
                cell_g = worksheet[f'G{row}']
                if cell_g.value:
                    cell_g.number_format = 'DD-MM-YYYY'
                # Fecha de vencimiento
                cell_h = worksheet[f'H{row}']
                if cell_h.value:
                    cell_h.number_format = 'DD-MM-YYYY'
                # Formatear números
                worksheet[f'I{row}'].number_format = '#,##0'
                worksheet[f'J{row}'].number_format = '#,##0'
                worksheet[f'K{row}'].number_format = '#,##0'
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'reporte_periodo_{fecha_inicio_obj.strftime("%Y%m%d")}_{fecha_fin_obj.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar el reporte: {str(e)}')
        return redirect('docs:reportes_periodo')


# ===========================
# Vistas para Procesar Planillas
# ===========================

import unicodedata

def limpiar_hoja(sheet):
    """Limpia las filas del encabezado y totales de una hoja de Excel"""
    # Buscar la fila que tenga exactamente 14 celdas no vacías (coincide con los encabezados)
    fila_encabezado = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        celdas_utiles = [cell for cell in row if cell is not None and str(cell).strip() != ""]
        if len(celdas_utiles) == 14:
            fila_encabezado = i
            break

    # Eliminar todas las filas anteriores al encabezado
    if fila_encabezado and fila_encabezado > 1:
        sheet.delete_rows(1, fila_encabezado - 1)

    # Buscar y eliminar la fila que contiene "TOTALES"
    for i, row in reversed(list(enumerate(sheet.iter_rows(values_only=True), start=1))):
        if any(str(cell).strip().upper() == "T O T A L E S" for cell in row if cell):
            sheet.delete_rows(i)
            break


def procesar_medios_pago(valor):
    """Convierte el nombre del medio de pago a su código"""
    valor_limpio = str(valor).strip()
    medios = {
        'Cuenta Prima': '01',
        'Cuenta prima': '01',
        'Cuenta Corriente / Vista': '01',
        'Cuenta RUT': '30',                        
        'Cuenta de Ahorro': '02',
        'Chequera Electrónica': '22',
        'Chequera Electronica': '22'
    }
    return medios.get(valor_limpio, None)


def procesar_codigo_banco(valor):
    """Convierte el nombre del banco a su código"""
    valor_limpio = str(valor).strip()
    bancos = {
        'BANCO BICE': '028',
        'BANCO CONSORCIO': '055',
        'BANCO DE CHILE-A EDWARDS-CITI': '001',
        'BCI-TBANC': '016',
        'BANCO ESTADO': '012',
        'BANCO FALABELLA': '051',
        'Banco Internacional': '009',
        'BANK BOSTON - ITAU': '039',
        'BANCO RIPLEY': '053',
        'BANCO SANTANDER - SANTIAGO': '037',
        'Banco Security': '049',
        'Caja de Compensacion Los Heroes': '729',
        'Banco Coopeuch': '672',
        'HSBC Bank Chile': '031',
        'BANCO SCOTIABANK': '014',
        'TENPO PREGAGO': '730',
        'CAJA DE COMPENSACION LOS ANDES-CUEN TAP': '732',
        'MERCADO PAGO': '875'
    }
    return bancos.get(valor_limpio, None)


def procesar_archivo_bancos_view(archivo):
    """Procesa archivo de bancos y retorna el workbook procesado"""
    workbook = openpyxl.load_workbook(archivo)
    sheet = workbook.active

    limpiar_hoja(sheet)  # Llama para limpiar encabezado
    
    workbook_nuevo = openpyxl.Workbook()
    sheet_nuevo = workbook_nuevo.active

    # Agregar encabezados
    sheet_nuevo.cell(row=1, column=1).value = "Nombre"
    sheet_nuevo.cell(row=1, column=2).value = "Detalle"
    sheet_nuevo.cell(row=1, column=3).value = "email"
    sheet_nuevo.cell(row=1, column=4).value = "Código Banco"
    sheet_nuevo.cell(row=1, column=5).value = "Medio de Pago"
    sheet_nuevo.cell(row=1, column=6).value = "Glosa"
    sheet_nuevo.cell(row=1, column=7).value = "Sueldo Líquido"

    sheet_nuevo.column_dimensions['F'].width = 25

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        valor_a = str(row[0].value) if row[0].value is not None else ""
        valor_b = str(row[1].value) if row[1].value is not None else ""
        sheet_nuevo.cell(row=i, column=1).value = valor_a + valor_b

        valor_e = row[4].value or ""
        valor_c = row[2].value or ""
        valor_d = row[3].value or ""
        detalle = f"{valor_e} {valor_c} {valor_d}"
        detalle = unicodedata.normalize(
            'NFKD', detalle).encode('ASCII', 'ignore').decode()
        detalle = detalle.replace("-", "").replace("\u00f1", "n")
        sheet_nuevo.cell(row=i, column=2).value = detalle

        sheet_nuevo.cell(row=i, column=3).value = ""

        codigo = {
            '1': '001', '01': '001', '12': '012', '14': '014', '16': '016',
            '28': '028', '37': '037', '39': '039', '51': '051', '53': '053',
            '55': '055', '672': '672', '729': '729', '730': '730', '732': '732'
        }.get(str(row[5].value), None)
        if codigo:
            sheet_nuevo.cell(row=i, column=4).value = codigo

        nombre_banco = row[6].value
        codigo_banco_procesado = procesar_codigo_banco(nombre_banco)
        if codigo_banco_procesado:
            sheet_nuevo.cell(row=i, column=4).value = codigo_banco_procesado

        tipo_medio_pago = row[10].value
        codigo_medio_pago_procesado = procesar_medios_pago(tipo_medio_pago)
        if codigo_medio_pago_procesado:
            sheet_nuevo.cell(row=i, column=5).value = codigo_medio_pago_procesado

        valor_glosa = str(row[9].value or "").replace(
            "-", "").replace(" ", "").replace(",", ".")
        try:
            valor_glosa_numerico = int(float(valor_glosa))
        except ValueError:
            valor_glosa_numerico = None
        celda = sheet_nuevo.cell(row=i, column=6)
        celda.value = valor_glosa_numerico if valor_glosa_numerico is not None else valor_glosa
        if valor_glosa_numerico is not None:
            celda.number_format = '0'

        sheet_nuevo.cell(row=i, column=7).value = row[13].value

    return workbook_nuevo


def procesar_archivo_vale_vista_view(archivo):
    """Procesa archivo de vale vista y retorna el workbook procesado"""
    workbook = openpyxl.load_workbook(archivo)
    sheet = workbook.active

    limpiar_hoja(sheet)  # Limpiar encabezados y totales

    workbook_nuevo_vv = openpyxl.Workbook()
    sheet_nuevo_vv = workbook_nuevo_vv.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=1):  # Empezar desde fila 2 para saltar el encabezado
        sheet_nuevo_vv.cell(row=i, column=1).value = "2"
        sheet_nuevo_vv.cell(
            row=i, column=2).value = row[0].value if row[0].value is not None else ""
        sheet_nuevo_vv.cell(
            row=i, column=3).value = row[1].value if row[1].value is not None else ""

        nombre = unicodedata.normalize('NFKD', str(
            row[4].value or "")).encode('ASCII', 'ignore').decode()
        nombre = nombre.replace("-", "").replace("\u00f1", "n")
        sheet_nuevo_vv.cell(row=i, column=4).value = nombre

        app = unicodedata.normalize('NFKD', str(row[2].value or "")).encode(
            'ASCII', 'ignore').decode()
        app = app.replace("-", "").replace("\u00f1", "n")
        sheet_nuevo_vv.cell(row=i, column=5).value = app

        apm = unicodedata.normalize('NFKD', str(row[3].value or "")).encode(
            'ASCII', 'ignore').decode()
        apm = apm.replace("-", "").replace("\u00f1", "n")
        sheet_nuevo_vv.cell(row=i, column=6).value = apm

        sheet_nuevo_vv.cell(row=i, column=7).value = "29"
        sheet_nuevo_vv.cell(row=i, column=8).value = "012"
        sheet_nuevo_vv.cell(row=i, column=9).value = "0"

        monto = row[13].value or ""
        sheet_nuevo_vv.cell(row=i, column=10).value = monto
        sheet_nuevo_vv.cell(row=i, column=17).value = monto
        sheet_nuevo_vv.cell(row=i, column=18).value = "M"

    return workbook_nuevo_vv


@login_required
def procesar_planillas(request):
    """Vista principal para procesar planillas de pago"""
    from .forms import ProcesarPlanillaForm
    
    if request.method == 'POST':
        form = ProcesarPlanillaForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                archivo = request.FILES['archivo']
                tipo_proceso = form.cleaned_data['tipo_proceso']
                
                # Procesar según el tipo
                if tipo_proceso == 'bancos':
                    workbook_procesado = procesar_archivo_bancos_view(archivo)
                    nombre_base = 'procesado_bancos'
                else:  # vale_vista
                    workbook_procesado = procesar_archivo_vale_vista_view(archivo)
                    nombre_base = 'procesado_vale_vista'
                
                # Crear respuesta con el archivo procesado
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                nombre_archivo = f"{nombre_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                
                workbook_procesado.save(response)
                
                messages.success(request, f'Archivo procesado exitosamente: {nombre_archivo}')
                return response
                
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
    else:
        form = ProcesarPlanillaForm()
    
    return render(request, 'docs/procesar_planillas.html', {'form': form})


# ===========================
# Vista de Ayuda
# ===========================
